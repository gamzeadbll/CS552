from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from openpyxl import Workbook, load_workbook
from scholarly import scholarly
import time
import random
import os


def fetch_author_details_by_scholar_id(scholar_id):
    try:
        author = scholarly.search_author_id(scholar_id)
        full_author = scholarly.fill(author)

        # Extract author details
        name = full_author.get("name", "N/A")
        affiliation = full_author.get("affiliation", "N/A")
        citations = full_author.get("citedby", "N/A")
        h_index = full_author.get("hindex", "N/A")
        i10_index = full_author.get("i10index", "N/A")
        citations_5y = full_author.get("citedby5y", "N/A")
        h_index_5y = full_author.get("hindex5y", "N/A")
        i10_index_5y = full_author.get("i10index5y", "N/A")
        publications = full_author.get("publications", [])
        total_publications = len(publications)

        return {
            "name": name,
            "affiliation": affiliation,
            "citations": citations,
            "h_index": h_index,
            "i10_index": i10_index,
            "citations_5y": citations_5y,
            "h_index_5y": h_index_5y,
            "i10_index_5y": i10_index_5y,
            "number_of_publications": total_publications,
        }
    except Exception as e:
        print(f"Error fetching details for Scholar ID {scholar_id}: {e}")
        return None


def scrape_google_scholar_authors(driver_path, university_name):
    service = Service(driver_path)
    driver = webdriver.Chrome(service=service)

    try:
        driver.get(f"https://scholar.google.com/citations?hl=en&view_op=search_authors&mauthors={university_name}&btnG=")

        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CLASS_NAME, "gs_ai"))
        )

        # Ensure the output directory exists
        os.makedirs("google_data", exist_ok=True)

        sanitized_name = university_name.replace("University", "").strip().replace(" ", "_")
        file_name = f"google_data/{sanitized_name}_authors.xlsx"

        if os.path.exists(file_name):
            wb = load_workbook(file_name)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Authors"
            ws.append([
                "Name", "Affiliation", "Citations", "Citations (Last 5 Years)",
                "h-index", "h-index (Last 5 Years)", "i10-index", "i10-index (Last 5 Years)",
                "Number of Publications", "Scholar ID", "Profile Link"
            ])

        page_counter = 0

        while True:
            authors = driver.find_elements(By.CLASS_NAME, "gs_ai")
            for author in authors:
                try:
                    profile_link_element = author.find_element(By.CLASS_NAME, "gs_ai_name").find_element(By.TAG_NAME, "a")
                    profile_link = profile_link_element.get_attribute("href")

                    scholar_id = profile_link.split("user=")[-1]

                    author_details = fetch_author_details_by_scholar_id(scholar_id)
                    if author_details:
                        ws.append([
                            author_details["name"],
                            author_details["affiliation"],
                            author_details["citations"],
                            author_details["citations_5y"],
                            author_details["h_index"],
                            author_details["h_index_5y"],
                            author_details["i10_index"],
                            author_details["i10_index_5y"],
                            author_details["number_of_publications"],
                            scholar_id,
                            profile_link,
                        ])
                        print(f"Added: {author_details['name']}")

                        wb.save(file_name)
                    else:
                        print(f"Failed to fetch details for Scholar ID: {scholar_id}")

                except Exception as e:
                    print(f"Error processing author: {e}")
                    continue

            try:
                # Locate the pagination container
                pagination_div = driver.find_element(By.ID, "gsc_authors_bottom_pag")

                # Find the "Next" button within the pagination container
                next_button = pagination_div.find_element(By.CSS_SELECTOR, ".gs_btnPR.gsc_pgn_pnx")
                is_clickable = next_button.get_attribute("onclick") is not None
                if is_clickable:
                    ActionChains(driver).move_to_element(next_button).perform()
                    next_button.click()
                    page_counter += 1
                    print(f"Processed page {page_counter}.")
                else:
                    print("The Next button is not clickable (disabled).")
                    break

                if page_counter % 10 == 0:
                    print("Pausing for 3 minutes after processing 10 pages...")
                    time.sleep(60)


                time.sleep(random.uniform(10, 30))
            except Exception as e:
                print(f"No more pages or error encountered: {e}")
                break

        print(f"Saved authors for {university_name} to {file_name}.")

    finally:
        driver.quit()