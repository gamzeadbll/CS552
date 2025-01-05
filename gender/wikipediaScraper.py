from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import unicodedata
import os


def normalize_name(name):
    # Remove any text within parentheses
    name = name.split("(")[0].strip()

    # Normalize Turkish characters
    translation_table = str.maketrans({
        "ç": "c", "Ç": "c",
        "ğ": "g", "Ğ": "g",
        "ı": "i", "İ": "i",
        "ö": "o", "Ö": "o",
        "ş": "s", "Ş": "s",
        "ü": "u", "Ü": "u",
    })
    normalized = name.translate(translation_table).lower()
    return ''.join(
        char for char in unicodedata.normalize('NFD', normalized)
        if unicodedata.category(char) != 'Mn'
    )


def write_to_excel(file_name, name, gender):
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Names"
        ws.append(["Name", "Gender"])  # Header row
        wb.save(file_name)

    wb = load_workbook(file_name)
    ws = wb.active
    ws.append([name, gender])
    wb.save(file_name)


def scrape_wikipedia_names(driver, excel_file):
    while True:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "mw-pages"))
        )

        mw_pages_div = driver.find_element(By.ID, "mw-pages")
        category_groups = mw_pages_div.find_elements(By.CLASS_NAME, "mw-category-group")

        for category_group in category_groups:
            ul_element = category_group.find_element(By.TAG_NAME, "ul")
            li_elements = ul_element.find_elements(By.TAG_NAME, "li")

            for li in li_elements:
                try:
                    link = li.find_element(By.TAG_NAME, "a")
                    name = link.text.strip()
                    normalized_name = normalize_name(name)
                    gender = "Female"
                    write_to_excel(excel_file, normalized_name, gender)
                    print(f"Added: {normalized_name} ({gender})")
                except Exception as e:
                    print(f"Error processing name: {e}")
                    continue

        try:
            # Find and click the "next page" link
            next_button = mw_pages_div.find_element(By.PARTIAL_LINK_TEXT, "next page")
            next_button.click()

            # Wait for the new page to load
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "mw-pages"))
            )
        except:
            print("No more pages available. Scraping complete.")
            break


if __name__ == "__main__":
    #female given names url = "https://en.wiktionary.org/wiki/Category:Turkish_female_given_names"
    url = "https://en.wiktionary.org/wiki/Category:Turkish_female_given_names"
    driver_path = "/Users/egeoruc/Downloads/chromedriver-mac-arm64/chromedriver"
    excel_file_name = "turkish_names_with_gender.xlsx"

    service = Service(driver_path)
    driver = webdriver.Chrome(service=service)

    try:
        driver.get(url)
        scrape_wikipedia_names(driver, excel_file_name)  # Pass only the driver and file name
    finally:
        driver.quit()