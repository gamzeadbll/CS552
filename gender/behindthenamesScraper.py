from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import os
import unicodedata


def normalize_name(name):
    # Normalize Turkish characters
    translation_table = str.maketrans({
        "ç": "c", "Ç": "c",
        "ğ": "g", "Ğ": "g",
        "ı": "i", "İ": "i",
        "ö": "o", "Ö": "o",
        "ş": "s", "Ş": "s",
        "ü": "u", "Ü": "u",
    })
    normalized = name.translate(translation_table).lower()
    return ''.join(
        char for char in unicodedata.normalize('NFD', normalized)
        if unicodedata.category(char) != 'Mn'
    )


def write_to_excel(file_name, name, gender):
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Names"
        ws.append(["Name", "Gender"])  # Header row
        wb.save(file_name)

    wb = load_workbook(file_name)
    ws = wb.active
    ws.append([name, gender])
    wb.save(file_name)


def scrape_behindthename(driver, excel_file):
    while True:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "body-inner"))
        )

        browsenames = driver.find_elements(By.CLASS_NAME, "browsename")

        for browsename in browsenames:
            try:
                name_element = browsename.find_element(By.CLASS_NAME, "listname").find_element(By.TAG_NAME, "a")
                name = name_element.text.strip()
                gender_element = browsename.find_element(By.CLASS_NAME, "listgender").find_element(By.TAG_NAME, "span")
                gender_short = gender_element.text.strip()

                if gender_short == "m":
                    gender = "Male"
                elif gender_short == "f":
                    gender = "Female"
                else:
                    gender = "Unisex"

                normalized_name = normalize_name(name)
                write_to_excel(excel_file, normalized_name, gender)

            except Exception as e:
                print(f"Error processing entry: {e}")
                continue

        try:
            pagination = driver.find_element(By.CLASS_NAME, "pagination")
            next_button = pagination.find_element(By.LINK_TEXT, "Next Page ►")
            next_button.click()
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "body-inner"))
            )
        except Exception:
            print("No more pages available. Scraping complete.")
            break


if __name__ == "__main__":
    url = "https://www.behindthename.com/submit/names/usage/turkish"
    driver_path = "/Users/egeoruc/Downloads/chromedriver-mac-arm64/chromedriver"
    excel_file_name = "behind_the_names.xlsx"

    service = Service(driver_path)
    driver = webdriver.Chrome(service=service)

    try:
        driver.get(url)
        scrape_behindthename(driver, excel_file_name)
    finally:
        driver.quit()